import torch.optim as optim
import torch.nn.functional as F
import openpyxl
import pickle
import torch.nn as nn
import pandas as pd
import pdb
import time
import math
from torch.optim.lr_scheduler import CosineAnnealingLR

from utils.utility import *
from utils.metric import *
from utils.construction import *
from tqdm import tqdm


### Trainer for 'SVM'
class SVM_Trainer:
    def __init__(self, args, device, network, train_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.test_loader = test_loader
    
    def train(self, i):
        for adjacency, feature, label in self.train_loader:
            self.network.fit(feature, label)
    
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        for adjacency, feature, label in self.test_loader:
            output = self.network.predict(feature)
            output = torch.FloatTensor(encode_onehot(output))

            loss_test = torch.tensor([0])
            accuracy_test = compute_accuracy(output, label)

            # print("Prediction Labels >")
            # print(output.max(1)[1])
            # print("Real Labels >")
            # print(label)

            # print(f"Test set results: loss_test: {loss_test.item():.3f} accuracy_test: {accuracy_test.item():.3f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"One vs One - Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'MLP'
class MLP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                # import pdb
                # pdb.set_trace()
                output = self.network.forward(feature) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                # import pdb
                # pdb.set_trace()
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            # print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GCN', 'GAT', 'GDC'
class GNN_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
        if args.model == 'gcn':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/gcn/"
        elif args.model == 'gdc':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/gdc/"
        elif args.model == 'adc':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/adc/"
        elif args.model == 'gat':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/gat/"
        elif args.model == 'sgformer':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/sgformer/"
        elif args.model == 'nodeformer':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/nodeformer/"
        elif args.model == 'difformer':
            self.checkpoint_path = "/root/simjy98/miccai2024/experiments/difformer/"

        self.timings = []
        self.starter, self.ender = torch.cuda.Event(enable_timing=True), torch.cuda.Event(enable_timing=True)
        
    def train(self, i):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        accuracy_best = 0.0

        pbar = tqdm(range(1, self.args.epochs + 1))
        pbar.set_description(f'FOLD-{i}')
        for epoch in pbar:
            self.network.train()

            # self.starter.record() ###

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                pbar.set_postfix({'loss': loss_train.item(), 'acc': accuracy_train.item()})
                
                # if epoch % 100 == 0:
                #     print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                # lt.append(loss_train.item())
                # at.append(accuracy_train.item())
            # self.ender.record()
            # torch.cuda.synchronize()
            # curr_time = self.starter.elapsed_time(self.ender)
            # self.timings.append(curr_time)
            # print(curr_time)
            # if epoch == 10:
            #     break

            self.network.eval()
            for adjacency, feature, label in self.test_loader:
                output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
            
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)
                
            if accuracy_test > accuracy_best:
                accuracy_best = accuracy_test
                torch.save(self.network.state_dict(), self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth')

        # mean = sum(self.timings)/len(self.timings)
        # vsum = 0
        # for val in self.timings:
        #     vsum = vsum + (val - mean)**2
        # variance = vsum / len(self.timings)
        # std = math.sqrt(variance)
        # print(mean, std)
        # exit(0)
        
    def test(self, i):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]
        
        self.network.load_state_dict(torch.load(self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth'))
        
        self.network.eval()
        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # print("Prediction Labels >")
            # print(output.max(1)[1])
            # print("Real Labels >")
            # print(label)
            
            # print(f"Test set results: loss_test: {loss_test.item():.3f} accuracy_test: {accuracy_test.item():.3f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GraphHeat'
class GraphHeat_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_sz):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_sz = adj_sz
        self.optimizer = optimizer
        
        self.checkpoint_path = "/root/simjy98/miccai2024/experiments/graphheat/"

        if args.use_t_local == 1: 
            self.t = torch.empty(adj_sz).fill_(2.)
        else:
            self.t = torch.tensor([2.])
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    def train(self, i):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        accuracy_best = 0.0
        
        pbar = tqdm(range(1, self.args.epochs + 1))
        pbar.set_description(f'FOLD-{i}')
        for epoch in pbar:
            self.network.train()

            for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                pbar.set_postfix({'loss': loss_train.item(), 'acc': accuracy_train.item()})
                
                # if epoch % 100 == 0:
                #     print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
                
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
                
                loss_test = F.nll_loss(output, label)
                accuracy_test = compute_accuracy(output, label)
                
            if accuracy_test > accuracy_best:
                accuracy_best = accuracy_test
                torch.save(self.network.state_dict(), self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth')

    def test(self, i):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.load_state_dict(torch.load(self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth'))

        self.network.eval()
        for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
            
            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            # print("Prediction Labels >")
            # print(output.max(1)[1])
            # print("Real Labels >")
            # print(label)
            
            # print(f"Test set results: loss_test: {loss_test.item():.3f} accuracy_test: {accuracy_test.item():.3f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'EXACT'
class Exact_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size
        self.optimizer = optimizer
        
        self.checkpoint_path = "/root/simjy98/miccai2024/experiments/exact/"

        if args.use_t_local == 1: # Local scale
            self.t = torch.empty(adj_size).fill_(2.)
        else: # Global scale
            self.t = torch.tensor([2.])
        
        self.t_lr = self.args.t_lr
        self.t_loss_threshold = self.args.t_loss_threshold
        self.t_alpha = self.args.t_alpha
        self.t_threshold = self.args.t_threshold
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_l = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)

        return self.t_alpha * torch.sum(t_l)

    def t_deriv(self):
        t_one = self.t_alpha * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def scale_update(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1)
        dl_ds = (torch.exp(output) - y_oh) / output.shape[0]
        
        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2))
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device)
            
        dl_dt += self.t_deriv() # Add regularizer on t
        now_lr = self.t_lr * dl_dt

        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold

        self.t = self.t - now_lr # Update t

        # if self.args.use_t_local == 1:
        #     print(f't:{self.t[0].item()}', end=' ')
        # else:
        #     print(f't:{self.t.item():.4f}', end=' ')

    ### Train
    def train(self, i):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        accuracy_best = 0.0
        
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.cell(row=1, column=1, value="epoch"+str(0))
        # for y in range(2, self.t.shape[0] + 2):
        #     ws.cell(row=1, column=y, value=2)
        # i = 2
        
        pbar = tqdm(range(1, self.args.epochs + 1))
        pbar.set_description(f'FOLD-{i}')
        for epoch in pbar:
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # if self.args.use_t_local == 1:
                #     with open('./variables/visualization/local.pickle', 'wb') as l:
                #         pickle.dump(heat_kernel, l)
                # else:
                #     with open('./variables/visualization/global.pickle', 'wb') as g:
                #         pickle.dump(heat_kernel, g)
                
                # with open('./variables/visualization/adjacency.pickle', 'wb') as a:
                #     pickle.dump(adjacency, a)
                # with open('./variables/visualization/laplacian.pickle', 'wb') as l:
                #     pickle.dump(laplacian, l)
                
                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                pbar.set_postfix({'loss': loss_train.item(), 'acc': accuracy_train.item()})
                
                # if epoch % 100 == 0:
                #     print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.3f} accuracy_train: {accuracy_train.item():.3f}")

                with torch.no_grad():
                    self.scale_update(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
            #     if epoch % 100 == 0 or epoch == self.args.epochs:
            #         s = self.t.detach().cpu().numpy()
            #         ws.cell(row=i, column=1, value="epoch"+str(epoch))
            #         for y in range(2, self.t.shape[0] + 2):
            #             ws.cell(row=i, column=y, value=s[y-2])
            #         i += 1
            
            # wb.save("/root/simjy98/tmi2023/experiments/scales/local"+str(fold)+".xlsx")
            self.network.eval()
            # with torch.no_grad():
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
                
                output_test = self.network(feature, heat_kernel) # Shape: (# of samples, # of labels)
        
                loss_test = F.nll_loss(output_test, label)
                accuracy_test = compute_accuracy(output_test, label)
                
            if accuracy_test > accuracy_best:
                accuracy_best = accuracy_test
                torch.save(self.network.state_dict(), self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth')

    ### Test
    def test(self, i):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.load_state_dict(torch.load(self.checkpoint_path + str(self.args.features) + '_' + str(self.args.labels) + '_' + str(i) + '_' + 'best_model.pth'))
        
        self.network.eval()
        for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)

            output = self.network(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)
            
            # print("Prediction Labels >")
            # print(output.max(1)[1])
            # print("Real Labels >")
            # print(label)
            
            # print(f"Test set results: loss_test: {loss_test.item():.3f} accuracy_test: {accuracy_test.item():.3f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        # return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts), label, output
        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)


### Trainer for 'LSAP'
class LSAP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size # 160
        self.optimizer = optimizer
        
        self.checkpoint_path = "/root/simjy98/miccai2024/experiments/lsap/"

        if args.use_t_local == 1: # Local scale 
            self.t = torch.empty(adj_size).fill_(2.) # (160)
        else: # Global scale 
            self.t = torch.tensor([2.]) # (1)
        
        self.t_lr = self.args.t_lr # 1
        self.t_loss_threshold = self.args.t_loss_threshold # 0.01
        self.t_lambda = self.args.t_alpha # 1
        self.t_threshold = self.args.t_threshold # 0.1
        
        self.m_che = self.args.m_chebyshev
        self.m_her = self.args.m_hermite 
        self.m_lag = self.args.m_laguerre 
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_loss_tmp = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)
        t_loss_final = self.t_lambda * torch.sum(t_loss_tmp) # λ|s|

        return t_loss_final

    def t_deriv(self):
        t_one = self.t_lambda * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def scale_update(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1) # (# sample, # label)

        dl_ds = (torch.exp(output) - y_oh) / output.shape[0] # (# sample, # label)

        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2)) # 160
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device) # 1

        dl_dt += self.t_deriv() # 160 / 1
        now_lr = self.t_lr * dl_dt # 160 / 1
        
        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold
        
        self.t = self.t - now_lr # Update t 

        # if self.args.use_t_local == 1:
        #     print(f't:{self.t[0].item()}', end=' ')
        # else:
        #     print(f't:{self.t.item():.4f}', end=' ')

    ### Train
    def train(self, i):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        accuracy_best = 0.0
        
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.cell(row=1, column=1, value="epoch"+str(0))
        # for y in range(2, self.t.shape[0] + 2):
        #     ws.cell(row=1, column=y, value=2)
        # i = 2

        pbar = tqdm(range(1, self.args.epochs + 1))
        pbar.set_description(f'FOLD-{i}')
        for epoch in pbar:
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian, P_n in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                if self.args.polynomial == 0: # Chebyshev
                    b = eigenvalue[:,eigenvalue.shape[1]-1]
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_chebyshev(P_n, laplacian, self.m_che, self.t, b, self.device)
                elif self.args.polynomial == 1: # Hermite
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_hermite(P_n, laplacian, self.m_her, self.t, self.device)
                elif self.args.polynomial == 2: # Laguerre
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_laguerre(P_n, laplacian, self.m_lag, self.t, self.device)

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
                # pdb.set_trace()
                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                pbar.set_postfix({'loss': loss_train.item(), 'acc': accuracy_train.item()})
                
                # if epoch % 100 == 0:
                #     print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                with torch.no_grad():
                    self.scale_update(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
                # s = self.t.detach().cpu().numpy()
                # ws.cell(row=i, column=1, value="epoch"+str(epoch))
                # for y in range(2, self.t.shape[0] + 2):
                #     ws.cell(row=i, column=y, value=s[y-2])
                # i += 1
            
            #wb.save("/home/user/ASAP_new/GC/our_s_"+ str(self.args.polynomial) +".xlsx")
            
            self.network.eval()
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian, P_n in self.test_loader:
                if self.args.polynomial == 0: # Chebyshev
                    b = eigenvalue[:,eigenvalue.shape[1]-1]
                    heat_kernel, _ = compute_heat_kernel_chebyshev(P_n, laplacian, self.m_che, self.t, b, self.device)
                elif self.args.polynomial == 1: # Hermite
                    heat_kernel, _ = compute_heat_kernel_hermite(P_n, laplacian, self.m_her, self.t, self.device)
                elif self.args.polynomial == 2: # Laguerre
                    heat_kernel, _ = compute_heat_kernel_laguerre(P_n, laplacian, self.m_lag, self.t, self.device)

                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
                
                loss_test = F.nll_loss(output, label) + self.t_loss()
                accuracy_test = compute_accuracy(output, label)
            
            if accuracy_test > accuracy_best:
                accuracy_best = accuracy_test
                torch.save(self.network.state_dict(), self.checkpoint_path + 
                           str(self.args.polynomial) + '_' + 
                           str(self.args.features) + '_' + 
                           str(self.args.labels) + '_' + 
                           str(i) + '_' + 'best_model.pth')
            
    ### Test
    def test(self, i):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.load_state_dict(torch.load(self.checkpoint_path + 
                                                str(self.args.polynomial) + '_' + 
                                                str(self.args.features) + '_' + 
                                                str(self.args.labels) + '_' + 
                                                str(i) + '_' + 'best_model.pth'))
        
        self.network.eval()     
        for adjacency, feature, label, eigenvalue, eigenvector, laplacian, P_n in self.test_loader:
            if self.args.polynomial == 0: # Chebyshev
                b = eigenvalue[:,eigenvalue.shape[1]-1]
                heat_kernel, _ = compute_heat_kernel_chebyshev(P_n, laplacian, self.m_che, self.t, b, self.device)
            elif self.args.polynomial == 1: # Hermite
                heat_kernel, _ = compute_heat_kernel_hermite(P_n, laplacian, self.m_her, self.t, self.device)
            elif self.args.polynomial == 2: # Laguerre
                heat_kernel, _ = compute_heat_kernel_laguerre(P_n, laplacian, self.m_lag, self.t, self.device)

            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            # pdb.set_trace()
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)
            
            # print("Prediction Labels >")
            # print(output.max(1)[1])
            # print("Real Labels >")
            # print(label)
            
            # print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)


### Trainer for 'Ours'
class OURS_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        self.device = device
        
        self.checkpoint_path = "/root/simjy98/miccai2024/experiments/ours/"
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)

    def t_loss(self):
        alpha = 1
        n_feats = self.network.n_feats
        t_loss_final_ = []
        for i in range(n_feats):
            t_abs = torch.abs(self.network.scale_per_feat[i])
            t_zero = torch.zeros_like(self.network.scale_per_feat[i])

            t_loss_tmp = torch.where(self.network.scale_per_feat[i] < self.args.t_loss_threshold, t_abs, t_zero)
            t_loss = torch.sum(t_loss_tmp) # λ|s|
            t_loss_final_.append(t_loss)
        
        t_loss_final = torch.mean(torch.stack(t_loss_final_))
        return alpha * t_loss_final
    
    def train(self, i):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        accuracy_best = 0.0
        
        pbar = tqdm(range(1, self.args.epochs + 1))
        pbar.set_description(f'FOLD-{i}')
        for epoch in pbar:
            self.network.train()
            

            for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                output = self.network.forward(feature, eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
                # output = self.network.inference_abl(feature, adjacency, eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
                
                loss_train_pr = F.nll_loss(output, label) + self.t_loss()

                accuracy_train = compute_accuracy(output, label)
                loss_train_pr.backward()

                pbar.set_postfix({'loss': loss_train_pr.item(), 't_loss': self.t_loss().item(), 'acc': accuracy_train.item()})
                
                self.optimizer.step() # Updates the parameters

            self.network.eval()
            with torch.no_grad():
                for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
                    output_test = self.network.inference(feature, eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
                    # output_test = self.network.inference_abl(feature, adjacency, eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
            
                    loss_test = F.nll_loss(output_test, label)
                    accuracy_test = compute_accuracy(output_test, label)
                    
                if accuracy_test > accuracy_best:
                    accuracy_best = accuracy_test
                    torch.save(self.network.state_dict(), 
                               self.checkpoint_path + 
                               str(self.args.features) + '_' +
                               str(self.args.labels) + '_' +
                               str(i) + '_' + 
                               'best_model.pth')

    def test(self, i):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.load_state_dict(torch.load(self.checkpoint_path + 
                                                str(self.args.features) + '_' + 
                                                str(self.args.labels) + '_' + 
                                                str(i) + '_' + 
                                                'best_model.pth'))
        # pdb.set_trace()
        self.network.eval()
        for adjacency, feature, label, eigenvalue, eigenvector, laplacian in self.test_loader:
            output = self.network.forward(feature, eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
            # output = self.network.inference_abl(feature, adjacency,eigenvalue, eigenvector, self.device) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Accuracy: {ac:.3f} Precision: {pr:.3f} Specificity: {sp:.3f} Sensitivity: {se:.3f} F1 score: {f1:.3f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None